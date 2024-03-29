# 起立歩行相分けプログラム(kerr et al., 2004)
# 筋電上位10個で正規化するコードも一応書いてる
# 床反力，位置，ロボット，重心速度，筋電をそれぞれファイルに入れる
# phaseフォルダにデータ保存，_ad.csvは床反力との同期後のファイル
#重心，重心速度も相分け済み
##左足から踏み込んでいる試行に使用

import pandas as pd
import numpy as np
import openpyxl
from matplotlib import pyplot
from tqdm import trange, tqdm
import os

#変更必要
#処理したいファイル名
filename = "n_10_"
#処理したい試行数名を書く
file_num = 4

print("処理中・・・", filename + str(file_num))

#各種サンプリング周波数
OPTITRACK_SAMPLING_RATE = 200 #[Hz]
GRF_SAMPLING_RATE = 100       #[Hz]
EMG_SAMPLING_RATE = 2000      #[Hz]
ROBOT_SAMPLING_RATE = 100     #[Hz]

#歩行相分けの床反力探索のための値 defalt=20
GRF_FORCE = 20 #[N]

#床反力探索したい区間[s] × サンプリング周波数 defalt=[2:5]
grfmin = int(2 * GRF_SAMPLING_RATE)
grfmax = int(7 * GRF_SAMPLING_RATE)
#踵探索したい区間[s]　× サンプリング周波数 defalt=[3.5:6.5]
heelmin = int(3.5 * OPTITRACK_SAMPLING_RATE)
heelmax = int(7 * OPTITRACK_SAMPLING_RATE)

#座位平均1秒間をとるスタート時間[s] defalt=2.5
sit_ave_start = 2.5

#起立開始探索区間[s] defalt=[4.5:7]
standmin = 4.5
standmax = 7

#フォルダなかったら作る（出力先）
os.makedirs("EXCEL/", exist_ok = True)
os.makedirs("phase/EMG/", exist_ok = True)
os.makedirs("phase/position/", exist_ok = True)
os.makedirs("phase/CoM_velocity/", exist_ok = True)
os.makedirs("phase/CoM/", exist_ok = True)
os.makedirs("phase/GRF/", exist_ok = True)
os.makedirs("phase/Robot/", exist_ok = True)
os.makedirs("tmp_tem/", exist_ok = True)

#処理開始
for i in tqdm(range(file_num, int(file_num + 1))):

    #ファイル読み込み
    df_grf = pd.read_csv("grf/" + filename + str(i) + ".csv")           #床反力
    df_position = pd.read_csv("position/" + filename + str(i) + ".csv") #位置データ
    df_comv = pd.read_csv("CoM_velocity/" + filename + str(i) + ".csv") #重心速度
    df_EMG = pd.read_csv("emg/" + filename + str(i) + ".csv")           #筋電
    df_robot = pd.read_csv("Robot/" + filename + str(i) + ".csv")       #ロボットデータ
    df_com = pd.read_csv("CoM/" + filename + str(i) + ".csv")           #重心

    #床反力オフセット
    R_sum_min = df_grf.R_Sum.min()
    L_sum_min = df_grf.L_Sum.min()
    df_grf["R_Sum"] = df_grf["R_Sum"] - R_sum_min
    df_grf["L_Sum"] = df_grf["L_Sum"] - L_sum_min

    #Excelデータ作成して同期時間を格納していく
    #Excelのファイルに新規のデータ作成
    wbname = "adjust_time_" + filename + str(i) + ".xlsx"
    wb = openpyxl.Workbook()
    wb.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")
    # ブックを取得
    book = openpyxl.load_workbook("Excel/adjust_time_" + filename + str(i) + ".xlsx")
    # シートを取得
    sheet = book["Sheet"]
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["C"].width = 20

    # セルへ書き込む
    sheet.cell(row = 1, column = 1).value = "相分け時間(s)"
    # 保存する
    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")

    #床反力グラフ作成
    fig1 = pyplot.figure()
    pyplot.plot(df_grf.L_Time, df_grf.L_Sum)
    pyplot.xlabel("Time(s)")
    pyplot.ylabel("GRF(N)")
    pyplot.xlim(0, )
    axis_array = np.arange(0, 21, step = 1)
    pyplot.xticks(axis_array)
    #Excelにグラフを貼り付け
    img1 = fig1.savefig("tmp_tem/tmp1.png") #tem_temフォルダ:貼り付ける画像を一時的に(temporarily)保存するフォルダ
    ximg1 = openpyxl.drawing.image.Image("tmp_tem/tmp1.png")
    sheet.add_image(ximg1, "E1")

    #踵グラフ作成
    fig2 = pyplot.figure()
    pyplot.plot(df_position["Time"], df_position["Body_LHeel_Position_Y"])
    pyplot.xlabel("Time(s)")
    pyplot.ylabel("Height(m)")
    pyplot.xlim(0, )
    axis_array = np.arange(0, 21, step = 1)
    pyplot.xticks(axis_array)
    #Excelにグラフを貼り付け
    img2 = fig2.savefig("tmp_tem/tmp2.png")
    ximg2 = openpyxl.drawing.image.Image("tmp_tem/tmp2.png")
    sheet.add_image(ximg2, "N1")
    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")


    ################################ 床反力との同期 ########################################
    #適当な区間で左脚床反力の最大値探索
    R_sum = df_grf.R_Sum
    L_sum = df_grf.L_Sum
    #床反力のデータ数カウント
    size_num = df_grf.L_Time.count()
    #探索区間×サンプリング周波数
    L_adjust = df_grf.L_Sum[grfmin:grfmax]
    #リスト化
    list_L = L_adjust.values.tolist()
    adtime = list_L.index(max(list_L))

    #探索区間の最小値×サンプリング周波数
    adnumber = grfmin + adtime + 1    #データ数 , adnumber:adjust number
    adjusted_time = (grfmin + adtime) / GRF_SAMPLING_RATE  #+ 0.1    #同期時間
    #床反力同期　+0.1秒後
    df_grf = df_grf.drop(range(adnumber + int(GRF_SAMPLING_RATE*0.1)))
    df_grf = df_grf.drop(columns=["L_Time"])
    #データ0~にする
    df_grf = df_grf.reset_index(drop = True)
    
    #時間オフセット用
    offset_times = []
    for step in range(len(df_grf)):
        offset_times.append(step / GRF_SAMPLING_RATE)
    #オフセット時間をデータフレームに追加
    df_grf.insert(0, "Time", offset_times, True)
    
    df_grf.to_csv("phase/GRF/" + filename + str(i) + "_ad.csv")

    #Excelファイルにメモ
    sheet.cell(row = 2, column = 1).value = "床反力同期時間"
    sheet.cell(row = 2, column = 2).value = adjusted_time
    book.save("Excel/adjust_time_" + filename + str(i) +".xlsx")

    #適当な区間で踵の最小値探索
    heelposition_z = df_position["Body_LHeel_Position_Y"]
    #データ数カウント
    psize_num = df_position["Time"].count()
    #探索区間の最小値×サンプリング周波数
    #探索区間×サンプリング周波数
    heel_adjust = heelposition_z[heelmin:heelmax]
    #リスト化
    list_heel = heel_adjust.values.tolist()
    #最小値探索
    adtime_position = list_heel.index(min(list_heel))

    adnumber_position = heelmin + adtime_position  #+ 1    #データ数
    adjusted_time_position = (heelmin + adtime_position) / OPTITRACK_SAMPLING_RATE + 0.1    #同期時間

    #Excelファイルにメモ
    sheet.cell(row=2,column=3).value = "位置データ同期時間"
    sheet.cell(row=2,column=4).value = adjusted_time_position
    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")

    #床反力と位置データの同期　+0.1秒後
    df_position = df_position.drop(range(adnumber_position + int(OPTITRACK_SAMPLING_RATE*0.1)))
    #データ0~にする
    df_position = df_position.reset_index(drop = True)

    #時間オフセット用
    offset_times = []
    for step in range(len(df_position)):
        offset_times.append(step / OPTITRACK_SAMPLING_RATE)
    #オフセット時間をデータフレームに追加
    df_position.insert(0, "Time", offset_times, True)
    
    df_position.to_csv("phase/position/" + filename + str(i) + "_ad.csv")

    #床反力と重心速度の同期　+0.1秒後
    df_comv = df_comv.drop(range(adnumber_position + int(OPTITRACK_SAMPLING_RATE*0.1)))
    #データ0~にする
    df_comv = df_comv.reset_index(drop = True)
    
    #時間オフセット用
    offset_times = []
    for step in range(len(df_comv)):
        offset_times.append(step / OPTITRACK_SAMPLING_RATE)

    #オフセット時間をデータフレームに追加
    df_comv.insert(0, "Time", offset_times, True)

    df_comv.to_csv("phase/CoM_velocity/" + filename + str(i) + "_ad.csv")

    
    #床反力と重心の同期　+0.1秒後
    df_com = df_com.drop(range(adnumber_position + int(OPTITRACK_SAMPLING_RATE*0.1)))
    #データ0~にする
    df_com = df_com.reset_index(drop = True)
    
    #時間オフセット用
    offset_times = []
    for step in range(len(df_com)):
        offset_times.append(step / OPTITRACK_SAMPLING_RATE)

    #オフセット時間をデータフレームに追加
    df_com.insert(0, "Time", offset_times, True)

    df_com.to_csv("phase/CoM/" + filename + str(i) + "_ad.csv")

    #床反力と筋電の同期　+0.1秒後
    df_EMG = df_EMG.drop(range(int(adnumber_position*EMG_SAMPLING_RATE / OPTITRACK_SAMPLING_RATE + EMG_SAMPLING_RATE*0.1)))
    
    #データ0~にする
    df_EMG = df_EMG.reset_index(drop = True)
    
    #時間オフセット用
    offset_times = []
    for step in range(len(df_EMG)):
        offset_times.append(step / EMG_SAMPLING_RATE)

    #オフセット時間をデータフレームに追加
    df_EMG.insert(0, "Time", offset_times, True)
    df_EMG.to_csv("phase/EMG/" + filename + str(i) + "_ad.csv")
    
    #床反力とRobotデータの同期
    df_robot = df_robot.drop(range(int(adnumber_position*ROBOT_SAMPLING_RATE / OPTITRACK_SAMPLING_RATE + ROBOT_SAMPLING_RATE*0.1)))
    #データ0~にする
    df_robot = df_robot.reset_index(drop = True)
   
    #時間オフセット用
    offset_times = []
    for step in range(len(df_robot)):
        offset_times.append(step / ROBOT_SAMPLING_RATE)
    #オフセット時間をデータフレームに追加
    df_robot.insert(0, "Time", offset_times, True)

    df_robot.to_csv("phase/Robot/" + filename + str(i) + "_ad.csv")
    
    #床反力との同期終了

    ############################# 相分け開始 ###############################

    #新しくファイル読みこみし直す
    df_grf = pd.read_csv("phase/GRF/" + filename + str(i) + "_ad.csv", index_col = 0)
    df_position = pd.read_csv("phase/position/" + filename + str(i) + "_ad.csv", index_col = 0)
    df_comv = pd.read_csv("phase/CoM_velocity/" + filename + str(i) + "_ad.csv", index_col = 0)
    df_com = pd.read_csv("phase/CoM/" + filename + str(i)+ "_ad.csv", index_col = 0)
    df_EMG = pd.read_csv("phase/EMG/" + filename + str(i) + "_ad.csv", index_col = 0)
    df_robot = pd.read_csv("phase/Robot/" + filename + str(i) + "_ad.csv", index_col = 0)
    
    #床反力合計
    Sum = df_grf.L_Sum + df_grf.R_Sum

    #床反力グラフ作成
    fig3 = pyplot.figure()
    pyplot.plot(df_grf.Time, df_grf.L_Sum)
    pyplot.plot(df_grf.Time, df_grf.R_Sum)
    pyplot.plot(df_grf.Time, Sum)
    pyplot.xlabel("Time(s)")
    pyplot.ylabel("GRF(N)")
    pyplot.xlim(0, 9)
    pyplot.ylim(0,)
    axis_array = np.arange(0, 21, step = 1)
    pyplot.xticks(axis_array)
    #Excelにグラフを貼り付け
    img3 = fig3.savefig("tmp_tem/tmp3.png")
    ximg3 = openpyxl.drawing.image.Image("tmp_tem/tmp3.png")
    sheet.add_image(ximg3, "E27")
    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")

    #踵グラフ作成
    fig4 = pyplot.figure()
    pyplot.plot(df_comv.Time, df_comv["CoM_y"])
    pyplot.xlabel("Time(s)")
    pyplot.ylabel("Velocity(m/s)")
    pyplot.xlim(0, 9)
    #pyplot.ylim(0,)
    axis_array = np.arange(0, 21, step = 1)
    pyplot.xticks(axis_array)
    #Excelにグラフを貼り付け
    img4 = fig4.savefig("tmp_tem/tmp4.png")
    ximg4 = openpyxl.drawing.image.Image("tmp_tem/tmp4.png")
    sheet.add_image(ximg4, "N27")
    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")
    

    ############################# 第1相 起立開始から臀部離床 ##################################

    #起立開始
    #床反力が座位静止時1秒間の平均値から2S.D.変化した瞬間
    
    #先頭からデータ抽出
    Sum_ave = Sum[int(GRF_SAMPLING_RATE*sit_ave_start):int(GRF_SAMPLING_RATE*(sit_ave_start + 1))]
    GRF_ave = Sum_ave.mean()
    GRF_2SD = 2*Sum_ave.std()
    #床反力
    GRF_stand_start = GRF_ave - GRF_2SD

    #起立開始の時間を探索，ループの回数はとりあえず適当にしてる
    for j in range(int(standmin*GRF_SAMPLING_RATE), int(standmax*GRF_SAMPLING_RATE)): 
        if Sum[j] <= GRF_stand_start.astype(float):

            stand_start = j
            stand_start_time = stand_start / GRF_SAMPLING_RATE
            #Excelファイルにメモ
            sheet.cell(row = 4, column = 1).value = "起立開始時間"
            sheet.cell(row = 4, column = 2).value = stand_start_time
            book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")
            break

    #臀部離床
    #床反力が最初に極大値を取った瞬間，最大値で検出してる．

    #400[N]以上の床反力探索
    for j in range(int(3*GRF_SAMPLING_RATE), df_grf.Time.count()): 
        if Sum[j] > 400:

            N_400 = j

            break
    
    #区間抽出してリスト化,最大値探す
    #床反力が400[N]超えた時点から最大値のインデックスを探索
    tmp_grf = Sum[N_400:N_400 + int(GRF_SAMPLING_RATE*0.7)]
    list_grf_sum = tmp_grf.values.tolist()
    Get_out_of_bed = N_400 + list_grf_sum.index(max(list_grf_sum)) #データ数
    Get_out_of_bed_time = Get_out_of_bed / GRF_SAMPLING_RATE #時間

    #Excelファイルにメモ
    sheet.cell(row = 5, column = 1).value = "臀部離床時間"
    sheet.cell(row = 5, column = 2).value = Get_out_of_bed_time
    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")

    df_grf_ad1 = df_grf.drop(range(int(stand_start)))
    df_position_ad1 = df_position.drop(range(int(stand_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_EMG_ad1 = df_EMG.drop(range(int(stand_start*EMG_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_robot_ad1 = df_robot.drop(range(int(stand_start*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_comv_ad1 = df_comv.drop(range(int(stand_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_com_ad1 = df_com.drop(range(int(stand_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
            
    #範囲を指定して，指定行以外削除
    df_grf_phase1 = df_grf_ad1.iloc[range(int(Get_out_of_bed - stand_start + 1)),:]
    df_grf_phase1.to_csv("phase/GRF/" + filename + str(i) + "_phase1.csv")
    df_position_phase1 = df_position_ad1.iloc[range(int((Get_out_of_bed - stand_start)*(OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_position_phase1.to_csv("phase/position/" + filename + str(i) + "_phase1.csv")
    df_EMG_phase1 = df_EMG_ad1.iloc[range(int((Get_out_of_bed - stand_start)*EMG_SAMPLING_RATE / GRF_SAMPLING_RATE)),:]
    df_EMG_phase1.to_csv("phase/EMG/" + filename + str(i) + "_phase1.csv")
    df_robot_phase1 = df_robot_ad1.iloc[range(int((Get_out_of_bed - stand_start)*(ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_robot_phase1.to_csv("phase/Robot/" + filename + str(i) + "_phase1.csv")
    df_comv_phase1 = df_comv_ad1.iloc[range(int((Get_out_of_bed - stand_start)*(OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_comv_phase1.to_csv("phase/CoM_velocity/" + filename + str(i) + "_phase1.csv")
    df_com_phase1 = df_com_ad1.iloc[range(int((Get_out_of_bed - stand_start)*(OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_com_phase1.to_csv("phase/CoM/" + filename + str(i) + "_phase1.csv")
    

    ########################## 第2相 臀部離床から起立終了 ###################################

    #高さ方向の最大重心速度を検出
    #最大値のインデックス探したかったけどNaNあったら探せなかったからとりあえず0で埋めた（重心速度の最大が0以下になるわけないから）
    df_comv_y = df_comv["CoM_y"]
    df_comv_y = df_comv_y.fillna(0)
    list_comv = df_comv_y.values.tolist()
    end_stand = list_comv.index(max(list_comv)) #データ数
    end_stand_time = end_stand / OPTITRACK_SAMPLING_RATE #時間

    #Excelファイルにメモ
    sheet.cell(row = 6, column = 1).value = "起立終了時間"
    sheet.cell(row = 6, column = 2).value = end_stand_time
    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")
    
    df_grf_ad2 = df_grf.drop(range(int(Get_out_of_bed)))
    df_position_ad2 = df_position.drop(range(int(Get_out_of_bed*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_EMG_ad2 = df_EMG.drop(range(int(Get_out_of_bed*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)))
    df_robot_ad2 = df_robot.drop(range(int(Get_out_of_bed*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_comv_ad2 = df_comv.drop (range(int(Get_out_of_bed*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_com_ad2 = df_com.drop(range(int(Get_out_of_bed*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
 
    #範囲を指定して，指定行以外削除
    df_grf_phase2 = df_grf_ad2.iloc[range(int((end_stand*GRF_SAMPLING_RATE / OPTITRACK_SAMPLING_RATE - Get_out_of_bed) + 1)),:]
    df_grf_phase2.to_csv("phase/GRF/" + filename + str(i) + "_phase2.csv")
    df_position_phase2 = df_position_ad2.iloc[range(int((end_stand - Get_out_of_bed*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_position_phase2.to_csv("phase/position/" + filename + str(i) + "_phase2.csv")
    df_EMG_phase2 = df_EMG_ad2.iloc[range(int(end_stand*EMG_SAMPLING_RATE/OPTITRACK_SAMPLING_RATE-Get_out_of_bed*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)),:]
    df_EMG_phase2.to_csv("phase/EMG/"+filename + str(i) + "_phase2.csv")
    df_robot_phase2 = df_robot_ad2.iloc[range(int((end_stand*ROBOT_SAMPLING_RATE / OPTITRACK_SAMPLING_RATE - Get_out_of_bed*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_robot_phase2.to_csv("phase/Robot/" + filename + str(i) + "_phase2.csv")
    df_comv_phase2 = df_comv_ad2.iloc[range(int((end_stand - Get_out_of_bed*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_comv_phase2.to_csv("phase/CoM_velocity/" + filename + str(i) + "_phase2.csv")
    df_com_phase2 = df_com_ad2.iloc[range(int((end_stand - Get_out_of_bed*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_com_phase2.to_csv("phase/CoM/" + filename + str(i) + "_phase2.csv")

  
    ############################ 第3相 歩行動作 4歩目まで相分け #########################################
    
    #床反力データで探索
    #床反力が，自身で定めた任意の値以下になることを検出することで相分けする．

    phase3_1_start = int(end_stand*GRF_SAMPLING_RATE / OPTITRACK_SAMPLING_RATE)

    #左脚から踏み出して，一旦左脚GRFが0になる時を探索
    #その後，左脚の接地時を探索
    #以下探索プログラムは同様の原理で組んでる
    for j in range(int(end_stand*GRF_SAMPLING_RATE / OPTITRACK_SAMPLING_RATE), df_grf.Time.count()):
        if df_grf.L_Sum[j] <= GRF_FORCE:
            
            left_grf_0 = j
            left_grf_0_time = j / GRF_SAMPLING_RATE
            break

    for j in range(left_grf_0,df_grf.Time.count()):
        if df_grf.L_Sum[j] > GRF_FORCE:
            
            phase3_1_end = j
            phase3_1_end_time = j / GRF_SAMPLING_RATE
            # セルへ書き込む
            sheet.cell(row = 7, column = 1).value = "3-1相終了時間"
            sheet.cell(row = 7, column = 2).value = phase3_1_end_time
            # 保存する
            book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")
            break


    df_grf_ad3_1 = df_grf.drop(range(phase3_1_start))
    df_position_ad3_1 = df_position.drop(range(int(phase3_1_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_EMG_ad3_1 = df_EMG.drop(range(int(phase3_1_start*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)))
    df_robot_ad3_1 = df_robot.drop(range(int(phase3_1_start*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_comv_ad3_1 = df_comv.drop(range(int(phase3_1_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_com_ad3_1 = df_com.drop(range(int(phase3_1_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))

    phase3_2_start = phase3_1_end 

    #範囲を指定して，指定行以外削除
    df_grf_phase3_1 = df_grf_ad3_1.iloc[range(phase3_2_start - phase3_1_start + 1),:]
    df_grf_phase3_1.to_csv("phase/GRF/" + filename + str(i) + "_phase3_1.csv")
    df_position_phase3_1 = df_position_ad3_1.iloc[range(int(((phase3_2_start - phase3_1_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_position_phase3_1.to_csv("phase/position/" + filename + str(i) + "_phase3_1.csv")
    df_EMG_phase3_1 = df_EMG_ad3_1.iloc[range(int((phase3_2_start-phase3_1_start)*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)),:]
    df_EMG_phase3_1.to_csv("phase/EMG/"+filename + str(i) + "_phase3_1.csv")
    df_robot_phase3_1 = df_robot_ad3_1.iloc[range(int(((phase3_2_start - phase3_1_start)*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_robot_phase3_1.to_csv("phase/Robot/" + filename + str(i) + "_phase3_1.csv")
    df_comv_phase3_1 = df_comv_ad3_1.iloc[range(int(((phase3_2_start - phase3_1_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_comv_phase3_1.to_csv("phase/CoM_velocity/" + filename + str(i) + "_phase3_1.csv")
    df_com_phase3_1 = df_com_ad3_1.iloc[range(int(((phase3_2_start - phase3_1_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_com_phase3_1.to_csv("phase/CoM/" + filename + str(i) + "_phase3_1.csv")


    for j in range(phase3_2_start,df_grf.Time.count()):
        if df_grf.R_Sum[j] <= GRF_FORCE:
            
            right_grf_0 = j
            break

    
    for j in range(right_grf_0,df_grf.Time.count()):
        if df_grf.R_Sum[j] > GRF_FORCE:
            
            phase3_2_end = j
            phase3_2_end_time = j / GRF_SAMPLING_RATE
            # セルへ書き込む
            sheet.cell(row = 8, column = 1).value = "3-2相終了時間"
            sheet.cell(row = 8, column = 2).value = phase3_2_end_time
            # 保存する
            book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")
            break

    
    df_grf_ad3_2 = df_grf.drop(range(phase3_2_start))
    df_position_ad3_2 = df_position.drop(range(int(phase3_2_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_EMG_ad3_2 = df_EMG.drop(range(int(phase3_2_start*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)))
    df_robot_ad3_2 = df_robot.drop(range(int(phase3_2_start*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_comv_ad3_2 = df_comv.drop(range(int(phase3_2_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_com_ad3_2 = df_com.drop(range(int(phase3_2_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))

    phase3_3_start = phase3_2_end

    #範囲を指定して，指定行以外削除
    df_grf_phase3_2 = df_grf_ad3_2.iloc[range(phase3_3_start - phase3_2_start + 1),:]
    df_grf_phase3_2.to_csv("phase/GRF/" + filename + str(i) + "_phase3_2.csv")
    df_position_phase3_2 = df_position_ad3_2.iloc[range(int(((phase3_3_start - phase3_2_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_position_phase3_2.to_csv("phase/position/" + filename + str(i) + "_phase3_2.csv")
    df_EMG_phase3_2 = df_EMG_ad3_2.iloc[range(int((phase3_3_start-phase3_2_start)*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)),:]
    df_EMG_phase3_2.to_csv("phase/EMG/"+filename + str(i) + "_phase3_2.csv")
    df_robot_phase3_2 = df_robot_ad3_2.iloc[range(int(((phase3_3_start - phase3_2_start)*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_robot_phase3_2.to_csv("phase/Robot/" + filename + str(i) + "_phase3_2.csv")
    df_comv_phase3_2 = df_comv_ad3_2.iloc[range(int(((phase3_3_start - phase3_2_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_comv_phase3_2.to_csv("phase/CoM_velocity/" + filename + str(i) + "_phase3_2.csv")
    df_com_phase3_2 = df_com_ad3_2.iloc[range(int(((phase3_3_start - phase3_2_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_com_phase3_2.to_csv("phase/CoM/" + filename + str(i) + "_phase3_2.csv")

    for j in range(phase3_3_start,df_grf.Time.count()):
        if df_grf.L_Sum[j] <= GRF_FORCE:
            
            left_grf_0 = j

            break

    for j in range(left_grf_0,df_grf.Time.count()):
        if df_grf.L_Sum[j] > GRF_FORCE:
            
            phase3_3_end = j
            phase3_3_end_time = j / GRF_SAMPLING_RATE
            # セルへ書き込む
            sheet.cell(row = 9, column = 1).value = "3-3相終了時間"
            sheet.cell(row = 9, column = 2).value = phase3_3_end_time
            # 保存する
            book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")
            break


    df_grf_ad3_3 = df_grf.drop(range(phase3_3_start))
    df_position_ad3_3 = df_position.drop(range(int(phase3_3_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_EMG_ad3_3 = df_EMG.drop(range(int(phase3_3_start*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)))
    df_robot_ad3_3 = df_robot.drop(range(int(phase3_3_start*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_comv_ad3_3 = df_comv.drop(range(int(phase3_3_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_com_ad3_3 = df_com.drop(range(int(phase3_3_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))

    phase3_4_start = phase3_3_end

    #範囲を指定して，指定行以外削除
    df_grf_phase3_3 = df_grf_ad3_3.iloc[range(phase3_4_start - phase3_3_start + 1),:]
    df_grf_phase3_3.to_csv("phase/GRF/" + filename + str(i) + "_phase3_3.csv")
    df_position_phase3_3 = df_position_ad3_3.iloc[range(int(((phase3_4_start - phase3_3_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_position_phase3_3.to_csv("phase/position/" + filename + str(i) + "_phase3_3.csv")
    df_EMG_phase3_3 = df_EMG_ad3_3.iloc[range(int((phase3_4_start-phase3_3_start)*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)),:]
    df_EMG_phase3_3.to_csv("phase/EMG/"+filename + str(i) + "_phase3_3.csv")
    df_robot_phase3_3 = df_robot_ad3_3.iloc[range(int(((phase3_4_start - phase3_3_start)*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_robot_phase3_3.to_csv("phase/Robot/" + filename + str(i) + "_phase3_3.csv")
    df_comv_phase3_3 = df_comv_ad3_3.iloc[range(int(((phase3_4_start - phase3_3_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_comv_phase3_3.to_csv("phase/CoM_velocity/" + filename + str(i) + "_phase3_3.csv")
    df_com_phase3_3 = df_com_ad3_3.iloc[range(int(((phase3_4_start - phase3_3_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_com_phase3_3.to_csv("phase/CoM/" + filename + str(i) + "_phase3_3.csv")

    for j in range(phase3_4_start,df_grf.Time.count()):
        if df_grf.R_Sum[j] <= GRF_FORCE:
            
            right_grf_0 = j

            break

    for j in range(right_grf_0,df_grf.Time.count()):
        if df_grf.R_Sum[j] > GRF_FORCE:
            
            phase3_4_end = j
            phase3_4_end_time = j / GRF_SAMPLING_RATE
            # セルへ書き込む
            sheet.cell(row = 10, column = 1).value = "3-4相終了時間"
            sheet.cell(row = 10, column = 2).value = phase3_4_end_time
            # 保存する
            book.save("Excel/adjust_time_" + filename + str(i) +".xlsx")
            break


    df_grf_ad3_4 = df_grf.drop(range(phase3_4_start))
    df_position_ad3_4 = df_position.drop(range(int(phase3_4_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_EMG_ad3_4 = df_EMG.drop(range(int(phase3_4_start*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)))
    df_robot_ad3_4 = df_robot.drop(range(int(phase3_4_start*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_comv_ad3_4 = df_comv.drop(range(int(phase3_4_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))
    df_com_ad3_4 =df_com.drop(range(int(phase3_4_start*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE)))

    phase3_5_start = phase3_4_end

    #範囲を指定して，指定行以外削除
    df_grf_phase3_4 = df_grf_ad3_4.iloc[range(phase3_5_start - phase3_4_start + 1),:]
    df_grf_phase3_4.to_csv("phase/GRF/" + filename + str(i) + "_phase3_4.csv")
    df_position_phase3_4 = df_position_ad3_4.iloc[range(int(((phase3_5_start - phase3_4_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_position_phase3_4.to_csv("phase/position/" + filename + str(i) + "_phase3_4.csv")
    df_EMG_phase3_4 = df_EMG_ad3_4.iloc[range(int((phase3_5_start-phase3_4_start)*EMG_SAMPLING_RATE/GRF_SAMPLING_RATE)),:]
    df_EMG_phase3_4.to_csv("phase/EMG/"+filename + str(i) + "_phase3_4.csv")
    df_robot_phase3_4 = df_robot_ad3_4.iloc[range(int(((phase3_5_start - phase3_4_start)*ROBOT_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_robot_phase3_4.to_csv("phase/Robot/" + filename + str(i) + "_phase3_4.csv")
    df_comv_phase3_4 = df_comv_ad3_4.iloc[range(int(((phase3_5_start - phase3_4_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_comv_phase3_4.to_csv("phase/CoM_velocity/" + filename + str(i) + "_phase3_4.csv")
    df_com_phase3_4 = df_com_ad3_4.iloc[range(int(((phase3_5_start - phase3_4_start)*OPTITRACK_SAMPLING_RATE / GRF_SAMPLING_RATE) + 1)),:]
    df_com_phase3_4.to_csv("phase/CoM/" + filename + str(i) + "_phase3_4.csv")

    ########################相時間#########################################

    #Excelファイルにメモ

    sheet.cell(row = 12, column = 1).value = "相継続時間(s)"

    sheet.cell(row = 14, column = 1).value = "1相"
    sheet.cell(row = 14, column = 2).value = Get_out_of_bed_time - stand_start_time
    phase1_time = Get_out_of_bed_time - stand_start_time
    
    sheet.cell(row = 15, column = 1).value = "2相"
    sheet.cell(row = 15, column = 2).value = end_stand_time - Get_out_of_bed_time
    phase2_time = end_stand_time - Get_out_of_bed_time

    sheet.cell(row = 16, column = 1).value = "3-1相"
    sheet.cell(row = 16, column = 2).value = phase3_1_end_time - end_stand_time
    phase3_1_time = phase3_1_end_time - end_stand_time

    sheet.cell(row = 17, column = 1).value = "3-2相"
    sheet.cell(row = 17, column = 2).value = phase3_2_end_time - phase3_1_end_time
    phase3_2_time = phase3_2_end_time -phase3_1_end_time

    sheet.cell(row = 18, column = 1).value = "3-3相"
    sheet.cell(row = 18, column = 2).value = phase3_3_end_time - phase3_2_end_time
    phase3_3_time = phase3_3_end_time - phase3_2_end_time

    sheet.cell(row =19, column = 1).value = "3-4相"
    sheet.cell(row =19, column =2).value = phase3_4_end_time - phase3_3_end_time
    phase3_4_time = phase3_4_end_time - phase3_3_end_time

    sheet.cell(row = 20, column = 1).value = "総時間"
    sheet.cell(row = 20, column = 2).value = phase1_time + phase2_time + phase3_1_time + phase3_2_time + phase3_3_time + phase3_4_time
    TOTAL_TIME = phase1_time + phase2_time + phase3_1_time + phase3_2_time + phase3_3_time + phase3_4_time

    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")

    ########################相割合#########################################

    #Excelファイルにメモ

    sheet.cell(row = 22, column = 1).value = "相割合(%)"
    
    sheet.cell(row = 24, column = 1).value = "1相"
    sheet.cell(row = 24, column = 2).value = (phase1_time / TOTAL_TIME) * 100
    sheet.cell(row = 24, column = 2).number_format = "0.00"

    sheet.cell(row = 25, column = 1).value = "2相"
    sheet.cell(row = 25, column = 2).value = (phase2_time / TOTAL_TIME) * 100
    sheet.cell(row = 25, column = 2).number_format = "0.00"

    sheet.cell(row = 26, column = 1).value = "3-1相"
    sheet.cell(row = 26, column = 2).value = (phase3_1_time / TOTAL_TIME) * 100
    sheet.cell(row = 26, column = 2).number_format = "0.00"

    sheet.cell(row = 27, column = 1).value = "3-2相"
    sheet.cell(row = 27, column = 2).value = (phase3_2_time / TOTAL_TIME) * 100
    sheet.cell(row = 27, column = 2).number_format = "0.00"

    sheet.cell(row = 28, column = 1).value = "3-3相"
    sheet.cell(row = 28, column = 2).value = (phase3_3_time / TOTAL_TIME) * 100
    sheet.cell(row = 28, column = 2).number_format = "0.00"

    sheet.cell(row = 29, column = 1).value = "3-4相"
    sheet.cell(row = 29, column = 2).value = (phase3_4_time / TOTAL_TIME) * 100
    sheet.cell(row = 29, column = 2).number_format = "0.00"

    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")

    ########################歩調#########################################

    #Excelファイルにメモ

    sheet.cell(row = 31, column = 1).value = "歩調(歩)"

    sheet.cell(row = 33, column = 1).value = "3-1相"
    sheet.cell(row = 33, column = 2).value = 60 / phase3_1_time
    sheet.cell(row = 33, column = 2).number_format = "0.00"

    sheet.cell(row = 34, column = 1).value = "3-2相"
    sheet.cell(row = 34, column = 2).value = 60 / phase3_2_time
    sheet.cell(row = 34, column = 2).number_format = "0.00"

    sheet.cell(row = 35, column = 1).value = "3-3相"
    sheet.cell(row = 35, column = 2).value = 60 / phase3_3_time
    sheet.cell(row = 35, column = 2).number_format = "0.00"

    sheet.cell(row = 36, column = 1).value = "3-4相"
    sheet.cell(row = 36, column = 2).value = 60 / phase3_4_time
    sheet.cell(row = 36, column = 2).number_format = "0.00"

    book.save("Excel/adjust_time_" + filename + str(i) + ".xlsx")