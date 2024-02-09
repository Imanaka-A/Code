#Tukey-Kramer検定
#analy実行後，使用する
#10%毎に検定する

import os
import openpyxl
import pandas as pd
import numpy as np
from statsmodels.stats.multicomp import pairwise_tukeyhsd

#項目，相，条件入力
phase = "phase1"
condition_A = "condition4"
condition_B = "condition7"
condition_C = "condition10"

os.makedirs("analysis/tukey/CoM_disp/x/" + phase + "/", exist_ok = True)
os.makedirs("analysis/tukey/CoM_disp/y/" + phase + "/", exist_ok = True)
os.makedirs("analysis/tukey/CoM_disp/z/" + phase + "/", exist_ok = True)

#CoM_disp_x
part = "CoM_disp_x"
wbname = "10_" + part + "_" + condition_A + condition_B + condition_C + ".xlsx"
wb = openpyxl.Workbook()
wb.save("analysis/tukey/CoM_disp/x/" + phase + "/" + wbname)

#ファイル名適宜変更
file_A =pd.read_excel("analysis/CoM_disp/x/" + phase + "/" + part + "_" + phase + "_" + condition_A + ".xlsx")
file_B = pd.read_excel("analysis/CoM_disp/x/" + phase + "/" + part + "_" + phase + "_" + condition_B + ".xlsx")
file_C = pd.read_excel("analysis/CoM_disp/x/" + phase + "/" + part + "_" + phase + "_" + condition_C + ".xlsx")

print("比較..", condition_A + condition_B + condition_C)

#Tukey実行
for i in range(0, 10):

    data_A = file_A.iloc[2:, i]
    data_A = data_A.dropna(how = "any", axis = 0)
    value_A = np.array(data_A)
    count_A = np.repeat([condition_A], len(value_A))
    
    data_B = file_B.iloc[2:, i]
    data_B = data_B.dropna(how = "any", axis = 0)
    value_B = np.array(data_B)
    count_B = np.repeat([condition_B], len(value_B))

    data_C = file_C.iloc[2:, i]
    data_C = data_C.dropna(how = "any", axis = 0)
    value_C = np.array(data_C)
    count_C = np.repeat([condition_C], len(value_C))

    value = np.hstack((value_A, value_B, value_C))
    count = np.hstack((count_A, count_B, count_C))

    tukey = pairwise_tukeyhsd(value, count)
    print("結果...", str((i)*10), "-", str((i + 1) * 10), "%")
    print(tukey)

    tukey = pd.DataFrame(data = tukey._results_table.data[1:], columns = tukey._results_table.data[0])
    
    with pd.ExcelWriter("analysis/tukey/CoM_disp/x/" + phase + "/" + wbname, mode = "a") as writer:
        tukey.to_excel(writer, sheet_name = str((i)*10) + "-" + str((i + 1) * 10))


#CoM_disp_y
part = "CoM_disp_y"
wbname = "10_" + part + "_" + condition_A + condition_B + condition_C + ".xlsx"
wb = openpyxl.Workbook()
wb.save("analysis/tukey/CoM_disp/y/" + phase + "/" + wbname)

#ファイル名適宜変更
file_A =pd.read_excel("analysis/CoM_disp/y/" + phase + "/" + part + "_" + phase + "_" + condition_A + ".xlsx")
file_B = pd.read_excel("analysis/CoM_disp/y/" + phase + "/" + part + "_" + phase + "_" + condition_B + ".xlsx")
file_C = pd.read_excel("analysis/CoM_disp/y/" + phase + "/" + part + "_" + phase + "_" + condition_C + ".xlsx")

print("比較..", condition_A + condition_B + condition_C)

#Tukey実行
for i in range(0, 10):

    data_A = file_A.iloc[2:, i]
    data_A = data_A.dropna(how = "any", axis = 0)
    value_A = np.array(data_A)
    count_A = np.repeat([condition_A], len(value_A))
    
    data_B = file_B.iloc[2:, i]
    data_B = data_B.dropna(how = "any", axis = 0)
    value_B = np.array(data_B)
    count_B = np.repeat([condition_B], len(value_B))

    data_C = file_C.iloc[2:, i]
    data_C = data_C.dropna(how = "any", axis = 0)
    value_C = np.array(data_C)
    count_C = np.repeat([condition_C], len(value_C))

    value = np.hstack((value_A, value_B, value_C))
    count = np.hstack((count_A, count_B, count_C))

    tukey = pairwise_tukeyhsd(value, count)
    print("結果...", str((i)*10), "-", str((i + 1) * 10), "%")
    print(tukey)

    tukey = pd.DataFrame(data = tukey._results_table.data[1:], columns = tukey._results_table.data[0])
    
    with pd.ExcelWriter("analysis/tukey/CoM_disp/y/" + phase + "/" + wbname, mode = "a") as writer:
        tukey.to_excel(writer, sheet_name = str((i)*10) + "-" + str((i + 1) * 10))


#CoM_disp_z
part = "CoM_disp_z"
wbname = "10_" + part + "_" + condition_A + condition_B + condition_C + ".xlsx"
wb = openpyxl.Workbook()
wb.save("analysis/tukey/CoM_disp/z/" + phase + "/" + wbname)

#ファイル名適宜変更
file_A =pd.read_excel("analysis/CoM_disp/z/" + phase + "/" + part + "_" + phase + "_" + condition_A + ".xlsx")
file_B = pd.read_excel("analysis/CoM_disp/z/" + phase + "/" + part + "_" + phase + "_" + condition_B + ".xlsx")
file_C = pd.read_excel("analysis/CoM_disp/z/" + phase + "/" + part + "_" + phase + "_" + condition_C + ".xlsx")

print("比較..", condition_A + condition_B + condition_C)

#Tukey実行
for i in range(0, 10):

    data_A = file_A.iloc[2:, i]
    data_A = data_A.dropna(how = "any", axis = 0)
    value_A = np.array(data_A)
    count_A = np.repeat([condition_A], len(value_A))
    
    data_B = file_B.iloc[2:, i]
    data_B = data_B.dropna(how = "any", axis = 0)
    value_B = np.array(data_B)
    count_B = np.repeat([condition_B], len(value_B))

    data_C = file_C.iloc[2:, i]
    data_C = data_C.dropna(how = "any", axis = 0)
    value_C = np.array(data_C)
    count_C = np.repeat([condition_C], len(value_C))

    value = np.hstack((value_A, value_B, value_C))
    count = np.hstack((count_A, count_B, count_C))

    tukey = pairwise_tukeyhsd(value, count)
    print("結果...", str((i)*10), "-", str((i + 1) * 10), "%")
    print(tukey)

    tukey = pd.DataFrame(data = tukey._results_table.data[1:], columns = tukey._results_table.data[0])
    
    with pd.ExcelWriter("analysis/tukey/CoM_disp/z/" + phase + "/" + wbname, mode = "a") as writer:
        tukey.to_excel(writer, sheet_name = str((i)*10) + "-" + str((i + 1) * 10))